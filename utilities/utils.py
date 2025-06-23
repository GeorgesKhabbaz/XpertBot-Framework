"""
utils.py
Contains logging setup and Excel data extraction functions.
"""

import inspect
import logging
import os
import datetime
from openpyxl import load_workbook


class AutomationLoggerClass:
    """
    Provides a method to initialize a logger with the caller's function name and
    write logs to a dated file inside the 'logs' folder.
    """
    @staticmethod
    def automation(log_level=logging.DEBUG):
        """
        Initializes a logger with a dynamic name based on the calling function's name,
        and logs to a dated log file in the 'logs/' directory.

        :param log_level: Logging level (e.g., logging.INFO, logging.DEBUG)
        :return: Configured logger object
        """
        # Get the caller's function name (helps identify which test/function the log came from)
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)

        # Format date for filename
        today = datetime.date.today()
        formatted_date = today.strftime("%Y-%m-%d")

        # Create log file path
        log_file_name = f"logs/automation_{formatted_date}.log"
        os.makedirs(os.path.dirname(log_file_name), exist_ok=True)

        # Configure file handler and formatter
        file_handler = logging.FileHandler(log_file_name, mode='a')
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)

        # Avoid adding multiple handlers
        if logger.hasHandlers():
            logger.handlers.clear()
        logger.addHandler(file_handler)

        return logger


# pylint: disable=C0301
TESTDATA_FOLDER = r"C:\Users\User\OneDrive - Lebanese University\Documents\InternshipXpertbotX6\XpertBotFramework\testdata"
DEFAULT_SHEET = "Sheet"


def get_newest_excel_file(folder_path=TESTDATA_FOLDER, sheet=DEFAULT_SHEET):
    """
    Loads the newest Excel file from a folder and extracts all rows into a list of dictionaries.

    :param folder_path: Path to the folder containing Excel files
    :param sheet: Optional sheet name (defaults to the first sheet)
    :return: List of dictionaries containing row data
    """
    # Get all Excel files in the folder
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    print(f"Files found in the folder: {files}")

    if not files:
        raise FileNotFoundError(
            f"No Excel files found in the specified folder: {folder_path}")

    # Sort files by creation time (most recent first)
    sorted_files = sorted(files, key=lambda f: os.path.getctime(
        os.path.join(folder_path, f)), reverse=True)
    print(f"Sorted files by creation time: {sorted_files}")

    # Pick the newest file
    latest_file = sorted_files[0]
    print(f"Newest file: {latest_file}")

    file_path = os.path.join(folder_path, latest_file)
    print(f"Full path of the newest file: {file_path}")

    # Load the workbook
    try:
        wb = load_workbook(filename=file_path)
    except Exception as e:
        raise FileNotFoundError(f"Failed to load the workbook: {e}") from e

    # Use the first sheet if none is specified
    if sheet is None:
        sheet = wb.sheetnames[0]

    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in workbook.")

    sh = wb[sheet]
    row_ct = sh.max_row
    col_ct = sh.max_column

    datalist = []

    # Read data row by row starting from the second row (assuming row 1 has headers)
    for row in range(2, row_ct + 1):
        row_data = {}
        for col in range(1, col_ct + 1):
            column_name = sh.cell(row=1, column=col).value
            column_value = sh.cell(row=row, column=col).value
            row_data[column_name] = column_value
        datalist.append(row_data)

    return datalist
